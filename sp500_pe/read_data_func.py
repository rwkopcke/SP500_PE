'''
   these are functions used by the update_data and display_data
   scripts
        
   access these values in other modules by
        import sp500_pe.read_data_func as rd
'''

import sys

from openpyxl import load_workbook
import openpyxl.utils.cell
import polars as pl
import polars.selectors as cs

import sp500_pe.helper_func as hp


def read_sp_date(file_addr, wksht_name,
                 date_keys, value_col_1, 
                 date_key_2, value_col_2,
                 column_names, include_prices= False):
    
    if date_keys is None:
        print('\n============================================')
        print(f'Date keys are {date_keys} for {wksht_name}')
        print(f'at:  {file_addr}')
        print('============================================\n')
        sys.exit()
    
    # load wksht from wkbk
    wksht = hp.find_wksht(file_addr, wksht_name)
    
    # fetch row for latest date and price
    key_row = hp.find_key_row(wksht, 'A', 1, date_keys)

    if (key_row == 0):
        print('\n============================================')
        print(f'Found no {date_keys} in {wksht_name}')
        print(f'at:  {file_addr}')
        print('============================================\n')
        sys.exit()
        
    name_date = hp.dt_str_to_date(
        wksht[f'{value_col_1}{key_row}'].value)
    
    if not include_prices:
        return [name_date, None]
        
    date_lst = []
    price_lst = []
    
    date_lst.append(name_date)
    name_date = name_date.date()  # value to return should be date()
    price_lst.append(wksht[f'{value_col_1}{key_row + 1}'].value)
    
    # fetch next date and price
    key_row = hp.find_key_row(wksht, 'A', key_row, date_key_2)
    
    if (key_row == 0):
        print('\n============================================')
        print(f'Found no {date_key_2} in {wksht_name}')
        print(f'at:  {file_addr}')
        print('============================================\n')
        sys.exit()
    
    date_lst.append(hp.dt_str_to_date(
        wksht[f'A{key_row - 2}'].value))
    price_lst.append(wksht[f'{value_col_2}{key_row -2}'].value)
    
    df = pl.DataFrame({
                column_names[0]: date_lst,
                column_names[1]: price_lst},
                schema= {column_names[0]: pl.Date, 
                            column_names[1]: pl.Float32})
    return [name_date, df]     
    

def read_sp_sheet(file_addr, wksht_name,
                  act_key, first_blk_col, last_blk_col,
                  skip_col, column_names):
    
    # load wksht from wkbk
    wksht = hp.find_wksht(file_addr, wksht_name)
    # fetch historical earnings data from wksht
    key_row = hp.find_key_row(wksht, 'A', 1, act_key)
    
    df = data_block_reader(wksht, key_row,
                           first_blk_col, last_blk_col,
                           skip_col, column_names)
    return df


def data_block_reader(wksht, key_row,
                      first_blk_col, last_blk_col,
                      skip_col, column_names):
    """
    This function uses read_wksht() and data_from_sheet()
    to return the block of data in a worksheet as a dataframe
    the first column is type datetime.date

    """
    
    # first data row to read. Follows key_wrd row.
    start_row = 1 + key_row
    # last data row to read. 
    stop_row = -1 + hp.find_key_row(wksht, 'A', 
                                    start_row, 
                                    [ None, 'Actuals'])
    
    # read a list of lists (rows)
    data = hp.data_from_sheet(wksht, first_blk_col, last_blk_col,
                              skip_col, start_row, stop_row)
    # iterate over rows to convert all dates to datetime.date
    # row[0]: datetime or str, '%m/%d/%Y' is first 'word' in str
    for row in data:
        row[0] = hp.dt_str_to_date(row[0])
    
    df = pl.DataFrame(data, schema=column_names, orient="row")\
                .cast({cs.float(): pl.Float32,
                       pl.Datetime: pl.Date})
    
    return df


def margin_reader(file_addr, wksht_name,
                  marg_key, first_blk_col, stop_col_key,
                  yr_qtr_name):
    """read the operating margin data for S&P 500 by year.
       one df for each date of projections

    Returns:
        data_df   : pl.DateFrame containing margin data
        name_date : date of projection
    """
    
    wksht = hp.find_wksht(file_addr, wksht_name)
    
    # find the rows with margin data
    start_row = hp.find_key_row(wksht, 'A', 1, marg_key)
    stop_row = start_row + 4
    
    # The label for the first col, which contains the qtr labels
    # is 'QTR'.  The other cols are named by year
    last_blk_col = -1 + hp.find_key_col(wksht, start_row, 
                                        2, stop_col_key)
    # block to collect: A,start_row  to  stop_col, start_row+4
    # collect data plus col headings for the block
    col_label = []
    for col_idx in range(1,last_blk_col):
        col = openpyxl.utils.cell.get_column_letter(col_idx)
        yr = str(wksht[f'{col}{start_row}'].value)
        col_label.append(yr)
    stop_blk_col =  openpyxl.utils.cell\
        .get_column_letter(last_blk_col)
    
    # list of lists for each row
    data = hp.data_from_sheet(wksht, first_blk_col, stop_blk_col, [],
                              start_row, stop_row)
    
    data_values = [
        [v for v in row]
        for row in data[1:] ]
    
    col_names = [str(item).split('*')[0] for item in data[0]]
   
    # build "tall" 2-col DF with 'year_qtr' and 'margin'
    df = pl.DataFrame(data_values, schema=col_names,
                      orient= 'row')\
                .with_columns(pl.col('QTR')
                              .map_elements(lambda x: x.split(' ')[0],
                                            return_dtype= str))\
                .cast({cs.float(): pl.Float32})\
                .unpivot(index= 'QTR', variable_name='year')
    df = df.with_columns(
                pl.struct(['QTR', 'year'])\
                    .map_elements(lambda x: 
                                  f"{x['year']}-{x['QTR']}",
                                  return_dtype= pl.String)\
                        .alias(yr_qtr_name))\
                    .drop(['year', 'QTR'])
    return df


def fred_reader(file_addr, rr_col_name, yr_qtr_name):
    wkbk = load_workbook(filename=file_addr,
                         read_only=True,
                         data_only=True)
    wksht = wkbk.active
    first_row = 12
    last_row = wksht.max_row
    
    data = hp.data_from_sheet(wksht, "A", "B", [],
                    first_row, last_row)
    
    df = pl.DataFrame(data, schema=['date', rr_col_name],
                      orient='row')\
           .with_columns(pl.col('date')
                        .map_batches(hp.date_to_year_qtr)
                        .alias(yr_qtr_name))\
           .cast({pl.Datetime: pl.Date,
                  cs.float(): pl.Float32})\
           .group_by('yr_qtr')\
           .agg([pl.all().sort_by('date').last()])\
           .sort(by= 'yr_qtr')\
           .drop('date')
    return df