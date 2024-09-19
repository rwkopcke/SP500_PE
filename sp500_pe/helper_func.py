'''
   these are functions used by the update_data and display_data
   scripts and by the read_data_func functions
        
   access these values in other modules by
        import sp500_pe.helper_func as hp
'''

from datetime import datetime

from openpyxl import load_workbook
import openpyxl.utils.cell
import polars as pl


def my_df_print(df):                                
    with pl.Config(
        tbl_cell_numeric_alignment="RIGHT",
        thousands_separator=",",
        float_precision=1
    ):
        print(df)
        
        
def dt_str_to_date(item):
    '''
        input either str or datetime obj
        return datetime object
        cast to datetime.date() in rd.data_block_reader()
    '''
    if isinstance(item, str):
        # fetch just the date component of str
        dt = item.split(" ")[0]
        dt = datetime.strptime(dt,'%m/%d/%Y')
    else:
        dt = item
    return dt
        

def string_to_date(series):
    '''
        receives pl.Series (col from df), str file names
        extracts the date string, "yyyy mm dd",
        returns the string to date object, as pl.Series
    '''
    # splits f_name into two two words,
    # the first and the rest
    # then splits the .xlsx from the rest
    return pl.Series([datetime.strptime((f_name.split(' ', 1)[1])
                                               .split('.', 1)[0], 
                                        '%Y %m %d').date()
                      for f_name in series])
    
    
def date_to_year_qtr(series):
    '''
        Receives pl.Series (col from df)
        Returns year_qtr string, yyyy-Qq, as pl.Series
    '''
    return pl.Series([f"{date.year}-Q{date_to_qtr(date)}"
            for date in series])
    
    
def date_to_qtr(date):
    '''
        Returns qtr number as string
    '''
    return f"{((date.month) - 1) // 3 + 1 }"


def is_quarter_4(series):
    return pl.Series([yq[-1] == '4'
                      for yq in series])
    
    
def yrqtr_to_yr(series):
    '''series of strings y-q in,
       series of strings y out
    '''
    return pl.Series([yq[:4]
                      for yq in series])
    
    
#======================================================================


def find_wksht(file_addr, wksht_name):
    """
    This helper function returns the worksheet
    for the file specified by the file address
    """

    wkbk = load_workbook(filename=file_addr,
                         read_only=True,
                         data_only=True)
    wksht = wkbk[wksht_name]
    return wksht


def find_key_row(wksht, search_col, start_row, key_values= None):

    # find cell containing the specified key,
    # return row number of the cell
    #   crawl down col A; return address of the first match
    
    # cap the number of rows to read
    max_to_read = 500
    
    row_number = start_row
    while row_number < max_to_read:
        item = wksht[f'{search_col}{row_number}'].value
        for key in key_values:
            if item_matches_key(item, key):
                 return row_number
        row_number += 1                                                     
    return 0


def item_matches_key(item, key):
    # all keys are either None or str
    if (key is None):
        return item is None
    if (isinstance(item, str)):
        return item == key
    return False


def find_key_col(wksht, search_row, start_col, key_value= None):
    
    # find cell containing the specified key,
    # return col letter of the cell
    #   crawl along search_row; return address of the first match
    
    # cap the number of rows to read
    max_to_read = 100
    
    col_numb = start_col
    while col_numb < max_to_read:
        # convert col number to letter
        col_ltr = openpyxl.utils.cell.get_column_letter(col_numb)
        item = wksht[f'{col_ltr}{search_row}'].value
        if item_matches_key(item, key_value):
            break
        col_numb += 1
    if col_numb == max_to_read:
        return 0                                                              
    return col_numb


def data_from_sheet(wksht, start_col, stop_col,
                    skip_col, start_row, stop_row):
    """
    This helper function fetches the block of data specified
    by first_col, start (row) and last_col, stop (row)
    in the specified worksheet
    """
    # rng references the block of data
    # the comprehension fetches the data over cols for each row
    rng = wksht[f'{start_col}{start_row}:{stop_col}{stop_row}']
    data = [[col_cell.value 
                for ind, col_cell in enumerate(row)
                if ind not in skip_col]
            for row in rng]
    return data
